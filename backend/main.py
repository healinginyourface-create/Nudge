from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from datetime import datetime
from openpyxl import load_workbook
import random
from pathlib import Path


app = FastAPI(title="Nudge API")


# ============================
# CORS 설정
# ============================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================
# Excel 활동 데이터
# ============================

DATA_PATH = Path(__file__).parent.parent / "data" / "activities.xlsx"


def load_activities():

    workbook = load_workbook(
        DATA_PATH,
        data_only=True
    )

    sheet = workbook["Activities"]

    activities = []

    headers = [
        cell.value
        for cell in sheet[1]
    ]

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row[0]:
            continue

        activity = dict(
            zip(headers, row)
        )

        # people
        if isinstance(activity.get("people"), str):

            activity["people"] = [
                person.strip()
                for person in activity["people"].split(",")
            ]

        # time_slots
        if isinstance(activity.get("time_slots"), str):

            activity["time_slots"] = [
                slot.strip()
                for slot in activity["time_slots"].split(",")
            ]

        # Excel에서 숫자가 float으로 읽히는 경우 정리
        if activity.get("cost") is not None:
            activity["cost"] = int(activity["cost"])

        if activity.get("duration") is not None:
            activity["duration"] = int(activity["duration"])

        # indoor
        if isinstance(activity.get("indoor"), str):

            activity["indoor"] = (
                activity["indoor"].lower() == "true"
            )

        activities.append(activity)

    workbook.close()

    return activities


# ============================
# 현재 시간대
# ============================

def get_current_time_slot():

    hour = datetime.now().hour

    if 6 <= hour < 12:
        return "morning"

    if 12 <= hour < 18:
        return "afternoon"

    if 18 <= hour < 22:
        return "evening"

    return "night"


# ============================
# 기본 API
# ============================

@app.get("/")
def root():

    return {
        "message": "Nudge API"
    }


# ============================
# 전체 활동 조회
# ============================

@app.get("/activities")
def get_activities():

    return load_activities()


# ============================
# 조건에 맞는 활동 전체 조회
# ============================

@app.get("/recommendations")
def get_recommendations(
    category: str | None = None,
    max_cost: int | None = None,
    people: str | None = None,
    indoor: bool | None = None
):

    activities = load_activities()

    recommendations = []

    for activity in activities:

        if category is not None:
            if activity["category"] != category:
                continue

        if max_cost is not None:
            if activity["cost"] > max_cost:
                continue

        if people is not None:
            if people not in activity["people"]:
                continue

        if indoor is not None:
            if activity["indoor"] != indoor:
                continue

        recommendations.append(activity)

    return recommendations


# ============================
# 활동 하나 추천
# ============================

@app.get("/recommend")
def recommend_activity(
    category: str | None = None,
    max_cost: int | None = None,
    people: str | None = None,
    indoor: bool | None = None
):

    activities = load_activities()

    candidates = []


    # ============================
    # 기본 조건 필터링
    # ============================

    for activity in activities:

        if category is not None:
            if activity["category"] != category:
                continue

        if max_cost is not None:
            if activity["cost"] > max_cost:
                continue

        if people is not None:
            if people not in activity["people"]:
                continue

        if indoor is not None:
            if activity["indoor"] != indoor:
                continue

        candidates.append(activity)


    # ============================
    # 조건에 맞는 활동이 없는 경우
    # ============================

    if not candidates:

        return {
            "message": "조건에 맞는 활동이 없습니다."
        }


    # ============================
    # 현재 시간대
    # ============================

    current_time_slot = get_current_time_slot()


    # ============================
    # 시간대 기반 점수 계산
    # ============================

    scored_candidates = []

    for activity in candidates:

        score = 0

        time_slots = activity.get(
            "time_slots",
            []
        )

        if current_time_slot in time_slots:

            score += 3


        scored_candidates.append(
            (activity, score)
        )


    # ============================
    # 최고 점수
    # ============================

    max_score = max(
        score
        for activity, score
        in scored_candidates
    )


    # ============================
    # 최고 점수 활동
    # ============================

    best_candidates = [

        activity

        for activity, score
        in scored_candidates

        if score == max_score
    ]


    # ============================
    # 랜덤 추천
    # ============================

    return random.choice(
        best_candidates
    )